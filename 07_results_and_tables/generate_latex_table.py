from openpyxl import load_workbook

# ======================
# 1. Load Excel
# ======================
wb = load_workbook("final.xlsx")
ws = wb.active
rows = list(ws.iter_rows())

header_lvl1 = [c.value for c in rows[0]]
header_lvl2 = [c.value for c in rows[1]]
data_rows   = rows[2:]

# ======================
# 2. Formatting helpers
# ======================
def format_value(val):
    try:
        return f"{float(val):.4f}"
    except (ValueError, TypeError):
        return "" if val is None else str(val)

def cell_to_tex(cell):
    val = format_value(cell.value)
    fill = cell.fill.fgColor
    if fill and fill.type == "rgb" and fill.rgb not in ("00000000", None):
        rgb = fill.rgb[-6:]
        return f"\\cellcolor[HTML]{{{rgb}}}{val}"
    return val

# ======================
# 3. Column format (equal width)
# ======================
MODEL_COL_WIDTH = "2.6cm"
SCORE_COL_WIDTH = "1.6cm"

col_format = []
for h in header_lvl2:
    if str(h).lower() == "score":
        col_format.append(
            f">{{\\raggedleft\\arraybackslash}}p{{{SCORE_COL_WIDTH}}}"
        )
    else:
        col_format.append(
            f">{{\\raggedright\\arraybackslash}}p{{{MODEL_COL_WIDTH}}}"
        )

col_format_str = "".join(col_format)

# ======================
# 4. Column headers
# ======================
group_cells = []
i = 0
while i < len(header_lvl1):
    if header_lvl1[i] is None:
        group_cells.append("\\multicolumn{1}{c}{}")
        i += 1
    else:
        group_cells.append(
            f"\\multicolumn{{2}}{{c}}{{{header_lvl1[i]}}}"
        )
        i += 2

group_header  = " & ".join(group_cells) + " \\\\"
second_header = " & ".join("" if v is None else str(v) for v in header_lvl2) + " \\\\"

# ======================
# 5. Table body with safe multirow
# ======================
ROWGROUP_SIZE = 6
SUBGROUP_SIZE = 3

body_lines = []
row_idx = 0

while row_idx < len(data_rows):
    for sub in range(2):
        for i in range(SUBGROUP_SIZE):
            row = data_rows[row_idx]
            cells = []

            # First column: big row group
            if sub == 0 and i == 0:
                cells.append(
                    f"\\multirow[t]{{{ROWGROUP_SIZE}}}{{*}}{{{row[0].value}}}"
                )
            else:
                cells.append("{}")

            # Second column: subgroup
            if i == 0:
                cells.append(
                    f"\\multirow[t]{{{SUBGROUP_SIZE}}}{{*}}{{{row[1].value}}}"
                )
            else:
                cells.append("{}")

            # Remaining score cells
            for cell in row[2:]:
                cells.append(cell_to_tex(cell))

            body_lines.append(" & ".join(cells) + " \\\\")
            row_idx += 1

# ======================
# 6. Final LaTeX table
# ======================
latex_table = "\n".join([
    "\\begin{table}[t]",
    "\\centering",
    "\\small",
    "\\setlength{\\tabcolsep}{4pt}",
    "\\renewcommand{\\arraystretch}{1.2}",
    "\\resizebox{\\textwidth}{!}{%",
    f"\\begin{{tabular}}{{{col_format_str}}}",
    "\\toprule",
    group_header,
    second_header,
    "\\midrule",
    "\n".join(body_lines),
    "\\bottomrule",
    "\\end{tabular}}",
    "\\end{table}",
])

with open("final_table.tex", "w", encoding="utf-8") as f:
    f.write(latex_table)

print("LaTeX table generated: final_table.tex (compilable)")
